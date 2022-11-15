# -*-coding: "utf-8"-*-

"""
# File:        data_process.py
# Author:      "liyi"
# CreateTime:  2022/6/20 15:12
# Version:     python 3.6
# Description: 将用例中的所有输入信号提取出来，组成输入矩阵；
#              将预期结果提取出来，组成预期结果矩阵；
"""
import copy
from collections import deque

import openpyxl
from copy import deepcopy
# from time import strftime
import re
import time


def is_contain_chinese(check_obj: str) -> bool:
    """
    判断字符串中是否包含中文
    :return: Y True, N False
    """
    for ch in check_obj:
        if u'\u4e00' <= ch <= u'\u9fff':
            return True
    return False

location_preprocess = 4
location_steps = 5
location_expect_result = 6
location_actual_result = 7

class DataProcess(object):
    def __init__(self, filepath):

        self.filepath = filepath

        # 初始化包含所有输入信号的矩阵
        self.InMatrix = deque()

        # 初始化包含所有输出信号的矩阵
        self.ExpMatrix = deque()

        # 初始化单例输入信号
        self.InSingle = deque()

        # 初始化单例输出信号
        self.ExpSingle = deque()

        # 初始化回归用例输入信号矩阵
        self.InMatrix_Rec = deque()

        # 初始化回归用力输出信号矩阵
        self.ExpMatrix_Rec = deque()

    def loadExcel(self):
        # 根据文件地址获取文档 workbook
        wb = openpyxl.load_workbook(self.filepath)
        # 获取sheet表格名称数组
        sheet_names = wb.sheetnames
        # 根据sheet名称获取sheet数据
        sheet_data = {}
        case_num = {}
        for item in sheet_names:
            sheet_data[item] = wb.get_sheet_by_name(item)
            # 获取用例个数
            case_num[item] = sheet_data[item].max_row - 2
        # print(sheet_data, case_num)
        # {'test1': <Worksheet "test1">, 'test2': <Worksheet "test1">}
        # {'test1': 2, 'test2': 2}
        # print(case_num)
        return sheet_data, case_num

    def _extractIn(self) -> deque:
        """
        提取表格中所有的输入信号，写入输入矩阵
        :return: 输入信号矩阵 InMatrix
        """
        # 初始化一个字典
        dict = {"pre_process": [], "steps": []}
        # dict['pre_process'] = []
        # dict['steps'] = []

        # 获取表格数据和用例个数
        sheet_data, case_num = self.loadExcel()

        # 将前置条件和操作步骤中的内容，依据换行符分割出来，获得信号的赋值动作
        keys = list(sheet_data.keys())
        # sheet个数
        len_ = len(keys)
        """
        self.InMatrix    |----------------------------------------------------------------------------------------------
        nrows = 1        |[{"pre_process":[操作1，操作2...], "steps":[步骤1，步骤2...]}, ... , 最后一个case的操作字典}
        ncols = sum_case |                                                                                  
                         |----------------------------------------------------------------------------------------------
        """
        for i in range(1, len_):
            for j in range(3, case_num[keys[i]] + 3):
                # 获取前置条件的信号设置
                cell_pre = sheet_data[keys[i]].cell(j, location_preprocess).value
                if cell_pre: # 判断前置条件不为空
                    dict["pre_process"] = cell_pre.split("\n")
                    for item in dict["pre_process"]:
                        if is_contain_chinese(item): # 去掉包含汉字的语句
                            dict["pre_process"].remove(item)
                else: # 若为空，赋空数组
                    dict["pre_process"] = []

                cell_step = sheet_data[keys[i]].cell(j, location_steps).value
                if cell_step:
                    dict["steps"] = cell_step.split("\n")
                    for item in dict["steps"]:
                        if is_contain_chinese(item):
                            dict["steps"].remove(item)
                else:
                    dict["steps"] = []
                self.InMatrix.append(deepcopy(dict))
                dict = {}
        # print(self.InMatrix)
        # 输出的结果格式如下：
        # deque([{'pre_process': ['MasterCylinderPressure=1', 'MasterCylinderQualifier=2'],
        #         'steps': ['SigGroup_27_ChkSm=3']},
        #        {'pre_process': ['MasterCylinderPressure=1', 'MasterCylinderQualifier=2'],
        #         'steps': ['SigGroup_27_ChkSm=3']},
        #        {'pre_process': ['MasterCylinderPressure=1', 'MasterCylinderQualifier=2'],
        #         'steps': ['SigGroup_27_ChkSm=3']},
        #        {'pre_process': ['MasterCylinderPressure=1', 'MasterCylinderQualifier=2'],
        #         'steps': ['SigGroup_27_ChkSm=3']}])
        return self.InMatrix

    def _extractIn_single(self, sheet_name: str, case_order: int) -> dict:
        """
        以 sheet 为维度提取用例输入
        :return:
        """
        # 定义一个空字典
        dict = {}

        # 根据 sheet_name 获取其中用例输入
        sheet_data, case_num = self.loadExcel()
        for k, v in sheet_data:
            if sheet_name == k:
                try:
                    cell_pre = sheet_data[k].cell(case_order + 2, location_preprocess).value
                    dict["pre_process"] = cell_pre.split('\n')
                    cell_steps = sheet_data[k].cell(case_order + 2, location_steps).value
                    dict["steps"] = cell_steps.split('\n')
                except IndexError as e:
                    print("error", "用例编号超出索引范围，请检查编号是否正确")
        return dict

    def _extractExp(self):
        """
        提取用例中的预期结果
        :return: 预期结果矩阵
        """
        dict = {}
        dict['expect'] = []

        # 获取表格数据和用例个数
        sheet_data, case_num = self.loadExcel()

        # 将前置条件和操作步骤中的内容，依据换行符分割出来，获得信号的赋值动作
        keys = list(sheet_data.keys())
        # sheet个数
        len_ = len(keys)
        # print(case_num)
        # 将预期结果中的内容，依据换行符分割出来
        for i in range(1, len_):
            # print(keys[i])
            for j in range(3, case_num[keys[i]] + 3):
                cell_pre = sheet_data[keys[i]].cell(j, location_expect_result).value
                # print(cell_pre)
                if len(cell_pre) == 1: # 单行
                    dict["expect"] = list(cell_pre)
                else:                  # 多行
                    dict["expect"] = cell_pre.split("\n")
                self.ExpMatrix.append(deepcopy(dict))
                dict = {}
        # print(self.ExpMatrix)
        # deque([{'expect': ['MasterCylinderPressure=1', 'MasterCylinderPressure=2', 'SigGroup_27_ChkSm=3']},
        #        {'expect': ['MasterCylinderPressure=1', 'MasterCylinderPressure=2', 'SigGroup_27_ChkSm=3']},
        #        {'expect': ['MasterCylinderPressure=1', 'MasterCylinderPressure=2', 'SigGroup_27_ChkSm=3']},
        #        {'expect': ['MasterCylinderPressure=1', 'MasterCylinderPressure=2', 'SigGroup_27_ChkSm=3']}])
        return self.ExpMatrix

    def _extractExp_single(self, sheet_name: str, case_order: int) -> dict:
        """
        提取单个用例的预期结果
        :return:
        """
        dict = {}
        sheet_data, case_num = self.loadExcel()
        for k, v in sheet_data:
            if k == sheet_name:
                try:
                    cell_exp = sheet_data[k].cell(case_order + 2, location_expect_result)
                    dict["expect"] = cell_exp.split('\n')
                except IndexError as e:
                    print("error", "用例编号超出索引范围，请检查编号是否正确")
        return dict

    def extract_sig_val(self, parseCSV: deque = None, parseDBC: dict = None) -> tuple:
        """
        从用例文档中提取出的输入和预期矩阵，经处理后，按照用例的维度，将输入信号及其数值、观测信号及其预期值
        存储到列表中
        :param in_operation: 输入矩阵，class deque
        :param expect_operation: 预期矩阵，class deque
        :return: 所有case的input list, output list
        """
        in_operation = self._extractIn()
        out_operation = self._extractExp()
        pre_process_list = []  # 前置条件 消息-信号-设定值 列表
        steps_list = []  # 测试步骤 消息-信号-设定值 列表
        input_list = []  # 整个输入 消息-信号-设定值 列表
        expect_list = []  # 预期结果 消息-信号-预期值 列表
        output_list = []  # 整个预期结果 消息-信号-预期值 列表

        # 输入为DBC解析内容
        if parseDBC:
            for item_case in in_operation:
                if not item_case['pre_process']:
                    pre_process_list = []
                else:
                    for item_pre_process in item_case['pre_process']:
                        item_split_process = item_pre_process.split('=')
                        content_process = item_split_process[0]
                        wait_time = self.get_wait_time(content_process)
                        # 获取XCP信号
                        if self.is_calibration_variable(content=content_process):
                            pre_process_list.append(item_split_process)

                        #  获取等待时间
                        elif wait_time:
                            pre_process_list.append(wait_time)
                        #  获取CAN信号和完整message信息
                        else:
                            item_split_process = self.get_can_signal(parsedbc=parseDBC, content=content_process, item_split_list=item_split_process)
                            pre_process_list.append(item_split_process)

                # 提取操作步骤的信号。若为空，则不做处理
                if not item_case['steps']:
                    steps_list = []
                else:
                    for item_steps in item_case['steps']:
                        item_split_steps = item_steps.split('=')
                        content_step = item_split_steps[0]
                        wait_time = self.get_wait_time(content_step)

                        # 获取XCP信号
                        if self.is_calibration_variable(content=content_step):
                            steps_list.append(item_split_steps)

                        #  获取等待时间
                        elif wait_time:
                            steps_list.append(wait_time)
                        #  获取CAN信号和完整message信息
                        else:
                            item_split_steps = self.get_can_signal(parsedbc=parseDBC, content=content_step,
                                                                     item_split_list=item_split_steps)
                            steps_list.append(item_split_steps)

                # input_list  [[[case1.preprocess1], [case1.preprocess2], ... , [case1.steps1], ... ],
                # [[case2.proprecess],[case2.steps]],...]
                input_list.append(pre_process_list + steps_list)
                pre_process_list.clear()
                steps_list.clear()

            for item_case in out_operation:
                for item_pre_expect in item_case['expect']:
                    item_split_expect = item_pre_expect.split('=')
                    content_expect = item_split_expect[0]
                    for item_dbc_name, item_dbc_info in parseDBC.items():
                        for item_single_dbc_info in item_dbc_info:
                            if content_expect == item_single_dbc_info['signal_name']:
                                # item_split_process: [message_name, signal_name, exp_val, dbc_name]
                                item_split_expect.insert(0, item_single_dbc_info['message_name'])
                                item_split_expect.append(item_dbc_name)
                                break
                        if len(item_split_expect) == 4:
                            expect_list.append(item_split_expect)
                            break
                    if len(item_split_expect) == 2:
                        if self.is_calibration_variable(content_expect):
                            expect_list.append(item_split_expect)
                        else:
                            raise RuntimeError(f"加载的dbc中无此信号: {content_expect}")
                output_list.append(copy.deepcopy(expect_list))
                expect_list.clear()

        # 输入为csv解析内容
        if parseCSV:
            for item_case in in_operation:
                for item_pre_process in item_case['pre_process']:
                    item_split_process = item_pre_process.split('=')
                    for item_csv_info in parseCSV:
                        if item_split_process[0] == item_csv_info['signal_name']:
                            # item_split_process: [message_name, signal_name, set_val]
                            item_split_process.insert(0, item_csv_info['message_name'])
                            break
                    if len(item_split_process) == 2:
                        raise RuntimeError(f"加载的dbc中无此信号: {item_split_process[0]}")
                    pre_process_list.append(item_split_process)
                for item_steps in item_case['steps']:
                    item_split_steps = item_steps.split('=')
                    for item_csv_info in parseCSV:
                        if item_split_steps[0] == item_csv_info['signal_name']:
                            item_split_steps.insert(0, item_csv_info['message_name'])
                            break
                    if len(item_split_steps) == 2:
                        raise RuntimeError(f"加载的dbc中无此信号: {item_split_steps[0]}")
                    steps_list.append(item_split_steps)
                input_list.append(pre_process_list + steps_list)
                pre_process_list.clear()
                steps_list.clear()

            for item_case in out_operation:
                for item_pre_expect in item_case['expect']:
                    item_split_expect = item_pre_expect.split('=')
                    for item_csv_info in parseCSV:
                        if item_split_expect[0] == item_csv_info['signal_name']:
                            # item_split_process: [message_name, signal_name, set_val]
                            item_split_expect.insert(0, item_csv_info['message_name'])
                            break
                    if len(item_split_expect) == 2:
                        raise RuntimeError(f"加载的dbc中无此信号: {item_split_expect[0]}")
                    expect_list.append(item_split_expect)
                output_list.append(copy.deepcopy(expect_list))
                expect_list.clear()

        return input_list, output_list

    def extract_sig_val_single(self, parseDBC: dict = None, parseCSV: deque = None, sheet_name: str = "",
                               case_order: int = 0):
        """
        获取单例的输入，预期输出信号
        :param parseDBC:解析后的dbc文件
        :param parseCSV: 解析后的csv文件
        :param sheet_name: 单例所在sheet
        :param case_order: 单例序号，用例中的编号，非表格行数
        :return:
        """
        in_operation = self._extractIn_single(sheet_name, case_order)  # dict
        out_operation = self._extractExp_single(sheet_name, case_order)  # dict

        input_pre_list = []
        input_steps_list = []
        output_exp_list = []
        # 初始化输出列表
        input_list = []
        output_list = []

        if parseDBC:
            for item_input_pre in in_operation['pre_process']:
                item_input_pre.split('=')
                for item_dbc_name, item_dbc_info in parseDBC.items():
                    for item_single_dbc_info in item_dbc_info:
                        if item_input_pre[0] == item_single_dbc_info['signal_name']:
                            item_input_pre.insert(0, item_single_dbc_info['message_name'])
                            item_input_pre.append(item_dbc_name)
                            item_input_pre.append(item_single_dbc_info['cycle_time'])
                            break
                    if len(item_input_pre) == 5:
                        break
                if len(item_input_pre) == 2:
                    if item_input_pre[0].startswith("P_"):
                        pass
                    else:
                        raise RuntimeError(f"加载的dbc中无此信号： {item_input_pre[0]}")
                input_pre_list.append(copy.deepcopy(item_input_pre))

            for item_input_steps in in_operation['steps']:
                item_input_steps.split('=')
                for item_dbc_name, item_dbc_info in parseDBC.items():
                    for item_single_dbc_info in item_dbc_info:
                        if item_input_steps[0] == item_single_dbc_info['signal_name']:
                            item_input_steps.insert(0, item_single_dbc_info['message_name'])
                            item_input_steps.append(item_dbc_name)
                            item_input_steps.append(item_single_dbc_info["cycle_time"])
                            break
                    if len(item_input_steps) == 5:
                        break
                if len(item_input_steps) == 2:
                    if item_input_steps[0].startswith("P_"):
                        pass
                    else:
                        raise RuntimeError(f"加载的dbc中无此信号： {item_input_steps[0]}")
                input_steps_list.append(copy.deepcopy(item_input_steps))

            input_list.append(input_pre_list + input_steps_list)

            for item_expect in out_operation['expect']:
                item_expect.split('=')
                for item_dbc_name, item_dbc_info in parseDBC.items():
                    for item_single_dbc_info in item_dbc_info:
                        if item_expect[0] == item_single_dbc_info['signal_name']:
                            item_expect.insert(0, item_single_dbc_info['message_name'])
                            item_expect.append(item_dbc_name)
                            break
                    if len(item_expect) == 4:
                        break

                if len(item_expect) == 2:
                    if item_expect[0].startswith("In_"):
                        pass
                    else:
                        raise RuntimeError(f"加载的dbc中无此信号： {item_expect[0]}")
                output_exp_list.append(copy.deepcopy(item_expect))
                output_list.append(output_exp_list)

        if parseCSV:
            for item_input_pre in in_operation['pre_process']:
                item_input_pre.split('=')
                for item_dbc_info in parseCSV:
                    if item_input_pre[0] == item_dbc_info['signal_name']:
                        item_input_pre.insert(0, item_dbc_info['message_name'])
                        break
                if len(item_input_pre) == 2:
                    raise RuntimeError(f"加载的dbc中无此信号： {item_input_pre[0]}")
                input_pre_list.append(copy.deepcopy(item_input_pre))

            for item_input_steps in in_operation['steps']:
                item_input_steps.split('=')
                for item_dbc_info in parseCSV:
                    if item_input_steps[0] == item_dbc_info['signal_name']:
                        item_input_steps.insert(0, item_dbc_info['message_name'])
                        break
                if len(item_input_steps) == 2:
                    raise RuntimeError(f"加载的dbc中无此信号： {item_input_steps[0]}")
                input_steps_list.append(copy.deepcopy(item_input_steps))

            input_list.append(input_pre_list + input_steps_list)

            for item_expect in out_operation['expect']:
                item_expect.split('=')
                for item_dbc_info in parseCSV:
                    if item_expect[0] == item_dbc_info['signal_name']:
                        item_expect.insert(0, item_dbc_info['message_name'])
                        break
                if len(item_expect) == 2:
                    raise RuntimeError(f"加载的dbc中无此信号： {item_expect[0]}")
                output_exp_list.append(copy.deepcopy(item_expect))
                output_list.append(output_exp_list)

        # input_list: [msg_name, sig_name, set_val, dbc_name, cycle_time]
        # output_list: [msg_name, sig_name, dbc_name]
        return input_list, output_list

    def get_wait_time(self, content)->list:
        if content.startswith("wait"):
            string_time = content.split(" ")[-1]
            return list(string_time.split("s")[0]) # 返回等待时长

    def is_calibration_variable(self, content)->bool:
        if content.startswith("P_") or \
            content.startswith("Out_") or \
            content.startswith("env_") or content.startswith("In_"):
            # print(str(content), " is a calibration\XCP object!")
            return True

    def get_can_signal(self, parsedbc: dict, content, item_split_list)->list:
        for item_dbc_name, item_dbc_info in parsedbc.items():
            for item_single_dbc_info in item_dbc_info:
                if content == item_single_dbc_info['signal_name']:
                    item_split_list.insert(0, item_single_dbc_info['message_name'])
                    item_split_list.append(item_dbc_name)
                    item_split_list.append(item_single_dbc_info["cycle_time"])
                    break
                else:
                    pass
            if len(item_split_list) == 5:
                break
        if len(item_split_list) == 5:
            return item_split_list
        else:
            raise RuntimeError(f"加载的dbc中无此信号: {content}")

# DBC 文件格式相关的参数
length_of_BO1 = 6  # BO_开头的行为message描述行，分割后可以形成长度为5的数组，BO_ 292 SP1_Info1_10ms: 32 FSD1
length_of_BO2 = 5
str_of_BO = 'BO_'  # 消息行开头
str_of_SG = 'SG_'  # 信号行开头
str_of_BA = 'BA_'  # 消息属性行开头
location_of_bo_id = 1  # 消息地址（id）所在位置
location_of_bo_message_name = 2  # 消息名所在位置
location_of_bo_dlc = 3  # 消息长度所在位置
location_of_bo_transmitter = 4  # 消息发送节点所在位置

# 信号行举例  SG_ TimeToRbump : 77|10@0+ (0.1,0) [0|102.3] ""  Vector_XXX
location_of_sg_type = 0
location_of_sg_name = 1
location_of_sg_factor = 4
location_of_sg_max_min = 5
location_of_sg_receiver = 7


# DBC load class
class DBCload(object):
    """
    解析处理单个DBC文件 .dbc
    """

    def __init__(self, dbc_name_in):
        self.dbc_fd = open(dbc_name_in, 'r')
        if self.dbc_fd.readable():
            self.num_of_bo = 0  # 预留
            self.num_of_sg = 0  # 预留
            self.dbc_list = []
            self.dbc_name = dbc_name_in
            self.dbc_cycle_time = {}
            self.dbc_send_type = {}
            # self.dbc_start_delay_time = {}
        else:
            print('DBC file load failed!')

    def parseDBC(self):
        """
        从DBC中获取消息和信号的属性
        消息： 名称  ID  发送方式   发送周期   重复发送次数  延迟发送时间
        信号： 名称  factor  offset  最大值  最小值
        发送周期：  周期发送       cyclic                           0
                  事件触发       spontaneous                      1
                  激活后循环     cyclicIfActive                   2
                  触发后延时     spontaneousWithDelay             3
                  触发后循环     cyclicAndSpontaneous             4
                  触发后延时循环  cyclicAndSpontaneousWithDelay    5
                  激活          ifActive                         6
        :return:
        """
        # 读取dbc文件
        line_list = self.dbc_fd.readlines()  # 逐行读取dbc，每行内容作为元素保存至列表中
        # print(line_list)
        dbc_name = self.dbc_name.split('/')[-1]
        dbc_txt_name = dbc_name.strip('.dbc')
        '''
        # 保存至txt文件中
        with open(r'../DBC' + time.strftime('%Y-%m-%d-%H-%M-%S') + dbc_txt_name + '.txt', 'w') as f:
            for item in line_list:
                f.write(item)
        '''
        # 提取消息内容
        for txt_line in line_list:
            txt_line_list = txt_line.split()
            # 查找消息属性描述行
            if len(txt_line_list) > 2 and txt_line_list[0] == str_of_BA:
                # 将 消息ID 和 发送周期 保存至dbc_cycle_time字典中
                if txt_line_list[1] == '"GenMsgCycleTime"':
                    self.dbc_cycle_time[txt_line_list[3]] = int(float(re.sub(';', '', txt_line_list[4])))
                # 将 消息ID 和 发送方式 保存至dbc_send_type字典中
                if txt_line_list[1] == '"GenMsgSendType"':
                    self.dbc_send_type[txt_line_list[3]] = int(re.sub(';', '', txt_line_list[4]))

        bo_list = []
        i = 0
        for i in range(len(line_list) - 1):
            txt_line_list = line_list[i].split()
            # 查找message描述行
            if (len(txt_line_list) == length_of_BO1 or len(txt_line_list) == length_of_BO2) and txt_line_list[
                0] == str_of_BO:
                # BO_ 661 FSD1_Info_FB_20ms : 32 Vector_XXX or BO_ 661 FSD1_Info_FB_20ms: 32 Vector_XXX
                # 记录消息名称、ID、发送周期、发送方式
                bo_dict = {'msg_name': re.sub(':', '', txt_line_list[location_of_bo_message_name]),
                           'msg_ID_DEC': int(txt_line_list[location_of_bo_id])}
                if str(bo_dict['msg_ID_DEC']) in self.dbc_cycle_time:
                    bo_dict['msg_cycle_time'] = self.dbc_cycle_time[str(bo_dict['msg_ID_DEC'])]
                else:
                    bo_dict['msg_cycle_time'] = 100000000
                if str(bo_dict['msg_ID_DEC']) in self.dbc_send_type:
                    bo_dict['msg_send_type'] = self.dbc_send_type[str(bo_dict['msg_ID_DEC'])]

                if line_list[i + 1] != "\n":  # 一条消息的信号簇以空行分段，故以此作为消息的分割符
                    i += 1
                    while len(line_list[i]) > 2 and line_list[i].split()[0] == str_of_SG:
                        sg_list = line_list[i].split()
                        sg_dict = {'message_name': bo_dict['msg_name'], 'signal_name': sg_list[location_of_sg_name],
                                   'msg_ID_DEC': bo_dict['msg_ID_DEC'], 'cycle_time': bo_dict['msg_cycle_time']}

                        if sg_list[2] != ':':
                            offset = 1
                        else:
                            offset = 0
                        end_of_factor = sg_list[location_of_sg_factor + offset].find(',')
                        end_of_offset = sg_list[location_of_sg_factor + offset].find(')')
                        sg_dict['factor'] = float(sg_list[location_of_sg_factor + offset][1:end_of_factor])
                        sg_dict['offset'] = float(
                            sg_list[location_of_sg_factor + offset][end_of_factor + 1:end_of_offset])
                        end_of_min = sg_list[location_of_sg_max_min + offset].find('|')
                        end_of_max = sg_list[location_of_sg_max_min + offset].find(']')
                        sg_dict['minimum'] = float(sg_list[location_of_sg_max_min + offset][1:end_of_min])
                        sg_dict['maximum'] = float(sg_list[location_of_sg_max_min + offset][end_of_min + 1:end_of_max])

                        bo_list.append(sg_dict)
                        i += 1

        dbc_text_name = dbc_txt_name.split('_')[-1]
        # print({dbc_text_name: bo_list})
        # 返回字典，如{"CHCAN1": [{"message_name": a, "signal_name": b,...},...]}
        # 返回dbc名称的原因是要配合canoe配置的通道，在收发信号时选择合适的通道
        return {dbc_text_name: bo_list}


# csv每行中不同条目的位置信息
location_sg = 0
location_ms = 1
location_Trans = 4
location_msId = 16
location_sig_cyc_time = 17
location_sig_timeout = 23


# csv load class
class csvload(object):
    def __init__(self, csv_name_in):
        self.csv_in = open(csv_name_in, 'r')
        if self.csv_in.readable():
            self.num_sg = 0
            self.text_lines = []
            self.table_sg_ms_msId_trans_sigCycTime_sigTimeOut = deque()
        else:
            raise FileExistsError('文件加载失败！')

    def parse_csv(self) -> deque:
        self.text_lines = self.csv_in.readlines()
        self.text_lines.pop(0)
        # print(self.text_lines)
        dict = {}
        de1 = deque()
        count = 0
        for item in self.text_lines:
            item.strip('\n')
            item = item.split(';')
            # print(item)
            dict['signal_name'] = item[location_sg]
            dict['message_name'] = item[location_ms]
            dict['message_Id'] = item[location_msId]
            dict['transmitter'] = item[location_Trans]
            dict['sigCycleTime'] = item[location_sig_cyc_time]
            # dict['sigTimeOut'] = item[location_sig_timeout]
            dict_new = deepcopy(dict)
            self.table_sg_ms_msId_trans_sigCycTime_sigTimeOut.append(dict_new)
        # print(self.table_sg_ms_msId_trans_sigCycTime_sigTimeOut)
        return self.table_sg_ms_msId_trans_sigCycTime_sigTimeOut